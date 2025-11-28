import os
import re
import time
import json
import shutil
import tempfile
import traceback
from datetime import datetime
from typing import List, Dict, Any, Optional
from urllib.parse import urlparse
import pandas as pd
import requests
import yaml
from dotenv import load_dotenv

# Selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# webdriver_manager (can be toggled/pinned via CHROMEDRIVER_VERSION)
from webdriver_manager.chrome import ChromeDriverManager

# Excel styling
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import base64
from botocore.exceptions import ClientError

PAGELOAD_TIMEOUT = 120


# ============================ CONFIG / ENV LOADER ============================
# Load local env overrides first (gitignored, optional)
# If you don't use .env.local this is harmless.
load_dotenv(dotenv_path=".env.local")
# Also load any top-level .env for backward compatibility
load_dotenv()

def _is_truthy(v: Optional[str]) -> bool:
    return str(v or "").strip().lower() in ("1", "true", "yes", "y", "on")

def load_yaml_as_env(path: str = "global_config.yaml") -> Dict[str, str]:
    """
    Read a structured YAML and flatten selected sections into environment variables.
    Example: runtime.headless -> RUNTIME_HEADLESS, driver.chromedriver_path -> DRIVER_CHROMEDRIVER_PATH
    """
    p = path
    try:
        if not os.path.exists(p):
            return {}
        with open(p, "r", encoding="utf-8") as fh:
            data = yaml.safe_load(fh) or {}
    except Exception:
        return {}

    flat: Dict[str, str] = {}
    for section, values in data.items():
        if isinstance(values, dict):
            for k, v in values.items():
                env_key = f"{section.upper()}_{k.upper()}"
                # Convert dict/list to JSON string, booleans/numbers to str
                if isinstance(v, (dict, list)):
                    flat[env_key] = json.dumps(v)
                elif v is None:
                    flat[env_key] = ""
                else:
                    flat[env_key] = str(v)
        else:
            flat[section.upper()] = str(values)
    return flat

def load_secrets_manager(secret_name: str) -> Dict[str, str]:
    """Fetch SecretString JSON from Secrets Manager and return it as a flat dict."""
    if not secret_name:
        return {}
    try:
        import boto3
        # Prefer env region, default to Seoul (your secret is there)
        region = (
            os.getenv("AWS_REGION")
            or os.getenv("AWS_DEFAULT_REGION")
            or "ap-northeast-2"
        )
        client = boto3.client("secretsmanager", region_name=region)
        resp = client.get_secret_value(SecretId=secret_name)
        secret_string = resp.get("SecretString")
        if not secret_string and resp.get("SecretBinary"):
            secret_string = base64.b64decode(resp["SecretBinary"]).decode("utf-8", errors="ignore")
        if not secret_string:
            return {}
        parsed = json.loads(secret_string)
        if isinstance(parsed, dict):
            return {k: str(v) for k, v in parsed.items()}
        return {}
    except Exception as e:
        print("Warning: could not load secrets from Secrets Manager:", e)
        return {}

# Apply YAML defaults into env (only if env var not already set)
yaml_env = load_yaml_as_env("src/config.yaml")
for k, v in yaml_env.items():
    if os.getenv(k) is None:
        os.environ[k] = v

# Optionally override with Secrets Manager
USE_SECRETS_MANAGER = _is_truthy(os.getenv("USE_SECRETS_MANAGER", "false"))
SECRET_NAME = os.getenv("SECRET_NAME", "").strip()
if USE_SECRETS_MANAGER and SECRET_NAME:
    secret_kv = load_secrets_manager(SECRET_NAME)
    for sk, sv in secret_kv.items():
        os.environ[sk] = sv  # overwrite YAML / .env values

# ============================ ENV / GLOBALS (from env after loader) ============================

def env_bool(name: str, default=False) -> bool:
    v = os.getenv(name, str(default)).strip().lower()
    if v is not None and str(v).strip() != "":
        return str(v)
    if default:
        print(f"Warning: {name} not set; falling back to default.")
    return default

LOGIN_URL         = os.getenv("LOGIN_URL", "https://service.xertica.cloud/dashboard")
OPSNOW_USERNAME  = env_with_fallback("OPSNOW_USERNAME",  "jaeyong.heo@bespinglobal.com")
OPSNOW_PASSWORD  = env_with_fallback("OPSNOW_PASSWORD",  "1qaz@WSX##")
XERTICA_USERNAME = env_with_fallback("XERTICA_USERNAME", OPSNOW_USERNAME)
XERTICA_PASSWORD = env_with_fallback("XERTICA_PASSWORD", OPSNOW_PASSWORD)


USERNAME = XERTICA_USERNAME
PASSWORD = XERTICA_PASSWORD
DEFAULT_SITE = "Xertica"
DEFAULT_COMPANY = "Xertica Clientes por reconocer"

EXCEL_FILE_PREFIX = os.getenv("REPORTING_EXCEL_FILE_PREFIX", "global_health_check_report")
SLACK_WEBHOOK_URL = os.getenv("SLACK_WEBHOOK_URL", "https://hooks.slack.com/services/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")
HEADLESS          = env_bool("RUNTIME_HEADLESS", env_bool("HEADLESS", False))
TIMEOUT           = int(os.getenv("RUNTIME_TIMEOUT", os.getenv("TIMEOUT", "30")))
RENDER_RETRY      = int(os.getenv("RUNTIME_RENDER_RETRY", os.getenv("RENDER_RETRY", "15")))
#LOCATOR_DEBUG     = env_bool("LOCATOR_DEBUG", False)
LOCATOR_DEBUG     = True
CHROMEDRIVER_VERSION = os.getenv("DRIVER_CHROMEDRIVER_VERSION", os.getenv("CHROMEDRIVER_VERSION", "142.0.7444.176")).strip()
CHROME_USER_DATA_DIR = os.getenv("DRIVER_CHROME_USER_DATA_DIR", os.getenv("CHROME_USER_DATA_DIR", "")).strip()

# Optional: fetch YAML from S3 instead of local file (set CONFIG_URI env if needed)
CONFIG_URI        = os.getenv("CONFIG_URI", "")  # e.g. s3://bucket/path/config.yaml
CONFIG_PATH_LOCAL = os.getenv("CONFIG_PATH_LOCAL", "xertica_config.yaml")

# Label keys for JS scan fallback
SERVER_LABEL_KEYS = ["total server", "total servers", "server", "servers", "서버", "총 서버"]

# ============================ FILE HELPERS ============================

driver: Optional[webdriver.Chrome] = None
wait:   Optional[WebDriverWait]    = None
TEMP_PROFILE_DIR: Optional[str]    = None

def ts() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

def ensure_dirs():
    os.makedirs("screenshots", exist_ok=True)
    os.makedirs("debug_html", exist_ok=True)
    os.makedirs("debug_json", exist_ok=True)

def safe_filename(name: str) -> str:
    if not isinstance(name, str):
        name = str(name)
    name = re.sub(r'[<>:"/\\|?*\n\r\t]+', "_", name)
    name = re.sub(r'__+', '_', name).strip("_ ")
    return name[:100] or "file"

def snap(name: str) -> str:
    ensure_dirs()
    path = os.path.join("screenshots", f"{safe_filename(name)}_{ts()}.png")
    try:
        if driver:
            driver.save_screenshot(path)
            print(f"Screenshot saved: {path}")
        else:
            # create an empty placeholder so caller has a path
            with open(path, "wb") as f:
                pass
            print(f"Driver not initialized; created placeholder screenshot: {path}")
    except Exception as e:
        print("Screenshot save failed:", e)

    # NEW: upload screenshot to S3 under SCREENSHOT_S3_PREFIX
    try:
        upload_to_s3(path, "SCREENSHOT_S3_PREFIX", "screenshots")
    except Exception as e:
        print(f"Screenshot S3 upload failed for {path}: {e}")

    return path


def dump_html(name: str) -> str:
    ensure_dirs()
    path = os.path.join("debug_html", f"{safe_filename(name)}_{ts()}.html")
    try:
        if driver:
            with open(path, "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print(f"HTML dumped: {path}")
        else:
            with open(path, "w", encoding="utf-8") as f:
                f.write("")
            print(f"Driver not initialized; created empty HTML dump: {path}")
    except Exception as e:
        print("HTML dump failed:", e)
    return path

def dump_json(name: str, obj: Any) -> str:
    ensure_dirs()
    path = os.path.join("debug_json", f"{safe_filename(name)}_{ts()}.json")
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)
        print(f"JSON dumped: {path}")
    except Exception as e:
        print("JSON dump failed:", e)
    return path



def upload_to_s3(local_path: str, prefix_env_var: str, default_prefix: str) -> None:
    """
    Upload a file to S3 using REPORT_S3_BUCKET and a prefix taken from an env var.
      - prefix_env_var: name of env var, e.g. "REPORT_S3_PREFIX" or "SCREENSHOT_S3_PREFIX"
      - default_prefix: used if that env var is not set
    """
    bucket = os.getenv("REPORT_S3_BUCKET", "").strip()
    if not bucket:
        print("REPORT_S3_BUCKET not set – skipping S3 upload.")
        return

    prefix = os.getenv(prefix_env_var, default_prefix) or ""
    prefix = prefix.strip().strip("/")

    file_name = os.path.basename(local_path)
    key = file_name if not prefix else f"{prefix}/{file_name}"

    try:
        import boto3
        s3 = boto3.client("s3")
        s3.upload_file(local_path, bucket, key)
        print(f"Uploaded {local_path} to s3://{bucket}/{key}")

        # If this is a screenshot, delete local file to save space
        if prefix_env_var == "SCREENSHOT_S3_PREFIX":
            try:
                os.remove(local_path)
                print(f"Deleted local screenshot after upload: {local_path}")
            except OSError as e:
                print(f"Could not delete local screenshot {local_path}: {e}")
    except Exception as e:
        print(f"Failed to upload {local_path} to S3: {e}")





# ============================ SLACK HELPERS ============================

def _slack_text(check_name: str, incident_msg: str, screenshot_path: Optional[str]) -> str:
    file_name = os.path.basename(screenshot_path) if screenshot_path else ""
    return (
        "OpsNow360 Health Check\n"
        f"- Incident Message : [{check_name}] : Fail - \"{incident_msg}\"\n"
        f"- Add Screenshot : {file_name}"
    )

def slack_notify(check_name: str, incident_msg: str, screenshot_path: Optional[str]):
    """
    Send a Slack message only if SLACK_WEBHOOK_URL is set.
    Log clearly whether Slack was sent / skipped / failed.
    """
    if not SLACK_WEBHOOK_URL:
        print(f"Slack: SKIP (SLACK_WEBHOOK_URL not set) for [{check_name}]")
        return

    try:
        file_name = os.path.basename(screenshot_path) if screenshot_path else ""
        text = (
            "OpsNow360 Health Check\n"
            f"- Incident : [{check_name}] : FAIL\n"
            f"- Message  : {incident_msg}\n"
            f"- Screenshot : {file_name}"
        )

        payload = {"text": text}
        print(f"Slack: sending alert for [{check_name}]...")
        r = requests.post(SLACK_WEBHOOK_URL, json=payload, timeout=10)

        if r.status_code < 300:
            print(f"Slack: OK (status {r.status_code}) for [{check_name}]")
        else:
            print(
                f"Slack: FAILED (status {r.status_code}) for [{check_name}] "
                f"body={r.text[:200]!r}"
            )

    except Exception as e:
        print(f"Slack: EXCEPTION for [{check_name}]: {e}")



# ============================ SELENIUM/SSO ============================

def create_driver():
    """
    Create a Chrome driver that supports:
    - ARM64 EC2 (installs chromium + chromium-driver)
    - AMD64 EC2 (installs Google Chrome + matching chromedriver)
    - Avoids webdriver-manager downloading wrong architecture binaries
    """
    global TEMP_PROFILE_DIR
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    import shutil
    import subprocess

    chrome_options = Options()
    if HEADLESS:
        chrome_options.add_argument("--headless=new")

    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-features=TranslateUI")

    # ---- USER DATA DIR FIX ----
    profile_dir = CHROME_USER_DATA_DIR or tempfile.mkdtemp(prefix="opsnow_chrome_")
    if not CHROME_USER_DATA_DIR:
        TEMP_PROFILE_DIR = profile_dir
    chrome_options.add_argument(f"--user-data-dir={profile_dir}")

    # ---- DETECT ARCH ----
    arch = subprocess.check_output(["uname", "-m"]).decode().strip()
    print(f"Detected architecture: {arch}")

    # ---- ARM64 PATH (Graviton) ----
    if arch in ("aarch64", "arm64"):

        print("ARM64 detected → using system chromium + chromium-driver")

        # Ensure chromium exists
        if not shutil.which("chromium") and not shutil.which("chromium-browser"):
            raise RuntimeError("Chromium is not installed inside the Docker image.")

        # Ensure chromedriver exists
        chromedriver_path = shutil.which("chromedriver")
        if not chromedriver_path:
            raise RuntimeError("chromium-driver is not installed inside the Docker image.")

        print(f"Using chromium driver at: {chromedriver_path}")
        driver = webdriver.Chrome(service=Service(chromedriver_path), options=chrome_options)

    # ---- AMD64 PATH ----
    else:
        print("AMD64 detected → using Google Chrome + matching chromedriver")

        try:
            chrome_bin = shutil.which("google-chrome") or shutil.which("chrome")
            if chrome_bin:
                chrome_options.binary_location = chrome_bin

            # Use webdriver-manager but ensure correct version
            if CHROMEDRIVER_VERSION:
                chromedriver_path = ChromeDriverManager(driver_version=CHROMEDRIVER_VERSION).install()
            else:
                chromedriver_path = ChromeDriverManager().install()

            print(f"Using chromedriver at: {chromedriver_path}")
            driver = webdriver.Chrome(service=Service(chromedriver_path), options=chrome_options)

        except Exception as e:
            raise RuntimeError(f"Failed to start Chrome on AMD64: {e}")

    driver.set_page_load_timeout(PAGELOAD_TIMEOUT)
    wait = WebDriverWait(driver, TIMEOUT)
    return driver, wait
    
    



def on_keycloak() -> bool:
    try:
        driver.find_element(By.NAME, "username")
        driver.find_element(By.NAME, "password")
        driver.find_element(By.ID, "kc-login")
        return True
    except NoSuchElementException:
        return False

def do_keycloak_login(current_url: Optional[str] = None,
                      username_override: Optional[str] = None,
                      password_override: Optional[str] = None):
    """
    Fill Keycloak credentials and submit.
    Priority:
      1) explicit username_override/password_override
      2) XERTICA_USERNAME/PASSWORD if current_url contains 'xertica'
      3) OPSNOW_USERNAME/PASSWORD
      4) global USERNAME/PASSWORD vars
    """
    if username_override is not None and password_override is not None:
        user = username_override
        pwd  = password_override
    else:
        if current_url and "xertica" in current_url.lower():
            user = os.getenv("XERTICA_USERNAME") or os.getenv("OPSNOW_USERNAME") or USERNAME
            pwd  = os.getenv("XERTICA_PASSWORD") or os.getenv("OPSNOW_PASSWORD") or PASSWORD
        else:
            user = os.getenv("OPSNOW_USERNAME") or USERNAME
            pwd  = os.getenv("OPSNOW_PASSWORD") or PASSWORD

    if not user or not pwd:
        raise ValueError("Keycloak credentials not found. "
                         "Set OPSNOW_USERNAME/OPSNOW_PASSWORD (and XERTICA_* for Xertica if needed).")

    try:
        u = wait.until(EC.presence_of_element_located((By.NAME, "username")))
        p = wait.until(EC.presence_of_element_located((By.NAME, "password")))
        b = wait.until(EC.element_to_be_clickable((By.ID, "kc-login")))
        u.clear(); u.send_keys(user)
        p.clear(); p.send_keys(pwd)
        b.click()
    except Exception:
        print("Exception while filling Keycloak login form:")
        traceback.print_exc()
        raise

def safe_do_keycloak_login(current_url: Optional[str] = None,
                           username_override: Optional[str] = None,
                           password_override: Optional[str] = None):
    try:
        do_keycloak_login(current_url=current_url,
                          username_override=username_override,
                          password_override=password_override)
    except Exception as e:
        print("safe_do_keycloak_login: failed:", e)
        traceback.print_exc()
        raise

def open_with_sso(url: str,
                  debug_name: str,
                  username_override: Optional[str] = None,
                  password_override: Optional[str] = None,
                  clear_cookies_before: bool = False):
    if clear_cookies_before:
        try:
            driver.delete_all_cookies()
        except Exception:
            pass
    try:
        driver.get(url)
    except Exception:
        print(f"Page load timeout for {url}")

    time.sleep(2)
    if on_keycloak():
        print(f"Keycloak detected — logging in for {url} …")
        safe_do_keycloak_login(current_url=url,
                               username_override=username_override,
                               password_override=password_override)
    try:
        WebDriverWait(driver, 60).until(
            lambda d: len(d.find_elements(By.CSS_SELECTOR,
                    "em.value, .value, .num, .number, .count, .am5-layer")) > 0
        )
    except TimeoutException:
        pass
    time.sleep(1)

def wait_non_empty_text(get_text, seconds=RENDER_RETRY) -> str:
    txt = ""
    for _ in range(seconds):
        try:
            txt = (get_text() or "").strip()
        except Exception:
            txt = ""
        if txt and txt != "0" and txt.lower() != "0ea":
            return txt
        time.sleep(1)
    return txt

def login_console():
    """Open LOGIN_URL and SSO in, waiting up to TIMEOUT for navigation."""
    if not LOGIN_URL:
        print("LOGIN_URL not set — skipping initial console login (will SSO per-page).")
        return
    print("Opening login page…", LOGIN_URL)

    try:
        driver.get(LOGIN_URL)
    except Exception:
        print(f"Page load timeout for {LOGIN_URL}")

    time.sleep(2)
    if on_keycloak():
        print("Logging in to console via Keycloak…")
        do_keycloak_login(current_url=LOGIN_URL)
    try:
        host = (urlparse(LOGIN_URL).hostname or "")
        if host:
            wait.until(EC.url_contains(host))
    except Exception:
        pass
    time.sleep(1)
    print("Login successful (or skipped).")


# ============================ JS FALLBACKS ============================

def js_scan_labels() -> List[Dict[str, str]]:
    js = r"""
    const blocks = Array.from(document.querySelectorAll(
      ".count-item, .summary, .card, .cards, [class*=count], [class*=summary], [class*=kpi]"
    ));
    const rows = [];
    const pull = (root) => {
      const labelEl = root.querySelector("p, .label, .title, h3, h4, dt, .name");
      const valueEl = root.querySelector("em.value, .value, .num, .number, .count, dd em.value");
      const label = (labelEl && labelEl.textContent ? labelEl.textContent.trim() : "");
      const value = (valueEl && valueEl.textContent ? valueEl.textContent.trim() : "");
      if (value) rows.push({label, value});
    };
    if (blocks.length) {
      blocks.forEach(pull);
    } else {
      const vals = Array.from(document.querySelectorAll("em.value, .value, .num, .number, .count"));
      vals.forEach(v => {
        let node = v, label = "";
        for (let i = 0; i < 5 && node; i++) {
          const l = node.querySelector?.("p, .label, .title, h3, h4, dt, .name");
          if (l && l.textContent) { label = l.textContent.trim(); break; }
          node = node.parentElement;
        }
        const value = v.textContent.trim();
        if (value) rows.push({label, value});
      });
    }
    return rows;
    """
    return driver.execute_script(js) or []

def pick_value_by_labels(rows: List[Dict[str, str]], label_keys: List[str]) -> str:
    for r in rows:
        lab = (r.get("label") or "").strip().lower()
        val = (r.get("value") or "").strip()
        if not val:
            continue
        for key in label_keys:
            if key.lower() in lab:
                return val
    for r in rows:
        lab = (r.get("label") or "").strip().lower()
        val = (r.get("value") or "").strip()
        if lab == "" and val and val.replace(",", "").isdigit():
            return val
    return ""

def js_find_ec2_near_aws() -> str:
    js = r"""
    const isVisible = (el) => {
      if (!el) return false;
      const st = getComputedStyle(el);
      return st && st.display !== 'none' && st.visibility !== 'hidden' && (el.offsetParent !== null || el.getClientRects().length);
    };
    const all = Array.from(document.querySelectorAll("span,div,button,a,li,p"));
    const ec2s = all.filter(el => (el.textContent || "").trim().toLowerCase() === "ec2" && isVisible(el));
    for (const ec2 of ec2s) {
      let node = ec2;
      for (let i = 0; i < 6 && node; i++) {
        const txt = (node.textContent || "").toLowerCase();
        if (txt.includes("aws")) return "FOUND";
        node = node.parentElement;
      }
    }
    const any = all.find(el => (el.textContent || "").trim().toLowerCase() === "ec2" && isVisible(el));
    return any ? "FOUND" : "";
    """
    return driver.execute_script(js) or ""

def js_find_mtd_cost() -> str:
    js = r"""
    const money = t => /\$\s*[\d,]+(\.\d+)?/.test((t||"").trim());
    const cards = Array.from(document.querySelectorAll("*"))
      .filter(el => /month\s*to\s*date\s*cost/i.test(el.textContent || ""));
    for (const card of cards) {
      const spans = card.querySelectorAll("span.currency-text, span[class*=currency], span[class*=-number], span[class*=value]");
      for (const s of spans) {
        const t = (s.textContent || "").trim();
        if (money(t)) return t;
      }
    }
    const spans = Array.from(document.querySelectorAll("span.currency-text, span[class*=currency], span[class*=-number], span[class*=value]"));
    for (const s of spans) {
      const t = (s.textContent || "").trim();
      if (money(t)) return t;
    }
    return "";
    """
    return driver.execute_script(js) or ""

def js_find_more_available_total() -> str:
    js = r"""
    const money = t => /\$\s*[\d,]+(\.\d+)?/.test((t||"").trim());
    const sections = Array.from(document.querySelectorAll("section, div, article"))
      .filter(el => /more\s+available\s+cost\s+savings/i.test(el.textContent || ""));
    for (const sec of sections) {
      const live = Array.from(sec.querySelectorAll("article"))
        .find(a => !/display\s*:\s*none/i.test(a.getAttribute("style") || ""));
      const root = live || sec;
      const values = root.querySelectorAll("span, div, p, b, strong");
      for (const v of values) {
        const t = (v.textContent || "").trim();
        if (money(t)) return t;
      }
    }
    return "";
    """
    return driver.execute_script(js) or ""

def js_find_ceikpi_grade_full() -> str:
    js = r"""
    const sections = Array.from(document.querySelectorAll("section,div,article"))
      .filter(el => /total\s*scores/i.test(el.textContent || ""));
    const parenRe = /^\([\d.,]+\)$/;
    for (const sec of sections) {
      const blocks = Array.from(sec.querySelectorAll("p,div,span"))
        .filter(el => /grade/i.test((el.textContent || "")));
      for (const b of blocks) {
        const spans = Array.from(b.querySelectorAll("span"));
        let grade = "";
        let paren = "";
        for (const s of spans) {
          const t = (s.textContent || "").trim();
          if (/grade/i.test(t)) grade = t;
          if (parenRe.test(t)) paren = t;
        }
        if (grade) {
          return (grade + (paren ? " " + paren : "")).trim();
        }
      }
    }
    return "";
    """
    return driver.execute_script(js) or ""


# ============================ CONFIG LOADING ============================

def maybe_fetch_config_from_s3(uri: str, local: str = "config.yaml") -> str:
    if not uri:
        return local
    import boto3
    assert uri.startswith("s3://")
    bucket, key = uri[5:].split("/", 1)
    s3 = boto3.client("s3")
    s3.download_file(bucket, key, local)
    return local

def load_config(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        raw = yaml.safe_load(f) or {}
    # expected: version, defaults, metadata_by_url, checks
    raw.setdefault("defaults", {})
    raw.setdefault("metadata_by_url", {})
    raw.setdefault("checks", [])
    return raw


# ============================ REPORTING ============================

EXCEL_COLUMNS = [
    "Site","Company","Service","Menu","URL","Check","Locator","Value","Status","Screenshot"
]

def make_result(meta: Dict[str, str], url: str, check_name: str,
                locator_used: str, value: str, status: str,
                screenshot_path: Optional[str] = None) -> Dict[str, str]:
    if status == "FAIL" and not screenshot_path:
        screenshot_path = snap(check_name.replace(" ", "_") + "_Fail")
    screenshot_name = os.path.basename(screenshot_path) if screenshot_path else ""
    return {
        "Site": meta.get("Site", "DEFAULT_SITE"),
        "Company": meta.get("Company", "DEFAULT_COMPANY"),
        "Service": meta.get("Service", ""),
        "Menu": meta.get("Menu", ""),
        "URL": url,
        "Check": check_name,
        "Locator": locator_used or "",
        "Value": value or "",
        "Status": status,
        "Screenshot": screenshot_name
    }

def save_report(rows: List[Dict[str, str]]) -> str:
    df = pd.DataFrame(rows)
    for col in EXCEL_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[EXCEL_COLUMNS]

    out = f"{EXCEL_FILE_PREFIX}_{ts()}.xlsx"
    df.to_excel(out, index=False)

    wb = load_workbook(out)
    ws = wb.active
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    status_col = header_map.get("Status")

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    if status_col:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=status_col)
            if cell.value == "PASS":
                cell.fill = green
            elif cell.value == "FAIL":
                cell.fill = red

    wb.save(out)
    print(f"\nReport saved locally: {out}")

    # Upload to S3 under REPORT_S3_PREFIX (e.g. reports/)
    upload_to_s3(out, "REPORT_S3_PREFIX", "reports")

    # If you truly want *only* S3 for reports, delete local file:
    try:
        os.remove(out)
        print(f"Local report deleted: {out}")
    except Exception as e:
        print(f"Could not delete local report {out}: {e}")

    return out


# ============================ XERTICA HELPERS ============================

def switch_company_to_force(target_text="Xertica Clientes por reconocer",
                            wait_after=2, timeout=10) -> bool:
    """
    Try hard to switch the topbar/company selector to the target.
    Returns True if verified in topbar text.
    """
    target = target_text.strip().lower()

    def read_topbar_text() -> str:
        sel_candidates = [
            "div.header__company","div.company-name","a.topbar-company","button.company-toggle","div.header .company",
            "div.bs-select-inline",".company-selector",".header-company"
        ]
        for s in sel_candidates:
            try:
                el = driver.find_element(By.CSS_SELECTOR, s)
                if el and el.text:
                    return el.text.strip().lower()
            except Exception:
                pass
        try:
            els = driver.find_elements(By.CSS_SELECTOR, "header *, nav *, div.topbar *, div.header *")
            for e in els:
                t = (e.text or "").strip().lower()
                if t and len(t) < 120 and ("xertica" in t or "*" in t or "transportes" in t):
                    return t
        except Exception:
            pass
        return ""

    def poll_for_verify(seconds=6):
        for _ in range(seconds):
            cur = read_topbar_text()
            if cur and target in cur:
                return True, cur
            time.sleep(1)
        return False, read_topbar_text()

    try:
        ok, cur = poll_for_verify(seconds=1)
        if ok:
            print("Company already selected:", cur)
            return True

        js_switch = r"""
        const target = arguments[0].trim().toLowerCase();
        const isVisible = el => !!el && getComputedStyle(el).display !== 'none' && getComputedStyle(el).visibility !== 'hidden' && (el.offsetParent !== null || el.getClientRects().length);
        const toggles = Array.from(document.querySelectorAll('a,button,div,span,p')).filter(n=>{
          try{
            const t=(n.innerText||'').trim().toLowerCase();
            return isVisible(n) && (t.includes('*') || /transportes|company|empresa|cliente|client|8091/i.test(t) || n.getAttribute('aria-haspopup')==='true' || n.getAttribute('role')==='button');
          }catch(e){return false;}
        });
        if(toggles.length){
          try{ toggles[0].scrollIntoView({block:'center',inline:'center'}); toggles[0].click(); }catch(e){}
        }
        const findAndClick = ()=>{
          const opts = Array.from(document.querySelectorAll('li,div,button,a,span,p'))
            .filter(n=> isVisible(n) && (n.innerText||'').trim().toLowerCase().includes(target));
          if(opts.length){
            try{ opts[0].scrollIntoView({block:'center',inline:'center'}); opts[0].click(); return true;}catch(e){}
            try{
              opts[0].dispatchEvent(new MouseEvent('mousedown',{bubbles:true}));
              opts[0].dispatchEvent(new MouseEvent('mouseup',{bubbles:true}));
              opts[0].dispatchEvent(new MouseEvent('click',{bubbles:true}));
              return true;
            }catch(e){}
          }
          return false;
        };
        if(findAndClick()) return true;
        const end = Date.now() + 2000;
        while(Date.now() < end){
          if(findAndClick()) return true;
        }
        return false;
        """
        clicked = bool(driver.execute_script(js_switch, target))
        print("Company selector JS click attempted:", clicked)

        ok, cur = poll_for_verify(seconds=4)
        if ok:
            print("Switched company (verified):", cur)
            return True

        print("Company switch not verified; attempting reload as last resort…")
        driver.refresh()
        time.sleep(1.5)
        ok, cur = poll_for_verify(seconds=6)
        if ok:
            print("Switched company after reload (verified):", cur)
            return True
        else:
            print("After reload — still not switched. topbar:", cur)
    except Exception as e:
        print("switch_company_to_force error:", e)
    return False

def select_only_xertica_option(option_text="Xertica Clientes por reconocer") -> bool:
    """Light attempt to click an option with visible text (scheduler inner panel)."""
    try:
        txt = option_text.strip().lower()
        opt_xpath = f"//*[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), {json.dumps(txt)})]"
        opts = driver.find_elements(By.XPATH, opt_xpath)
        for o in opts:
            if o.is_displayed():
                try:
                    o.click()
                    return True
                except Exception:
                    try:
                        driver.execute_script("arguments[0].click();", o)
                        return True
                    except Exception:
                        pass
        return False
    except Exception as e:
        print("select_only_xertica_option error:", e)
        return False


# ============================ CHECK DISPATCHER ============================

def run_one_check(check: Dict[str, Any], cfg: Dict[str, Any]) -> Dict[str, str]:
    start_time = time.time()
    MAX_CHECK_TIME = 180

    url = check["url"]
    name = check["name"]
    ctype    = (check.get("type") or "value_required").lower()
    locators = check.get("locators", [])
    js_fb    = check.get("js_fallback") or {}

    # --- Xertica Total Server locator override ---
    # If this is the "Total Server" check on the Xertica dashboard,
    # force known-good locators (label-based + index fallback).
    if "asset.xertica.cloud" in url and "Total Server" in name:
        locators = [
            {
                "kind": "xpath",
                "value": (
                    "//span[contains(@class,'key') and "
                    "contains(normalize-space(.), 'Total Server')]"
                    "/ancestor::dt[contains(@class,'count-item')]"
                    "//em[contains(@class,'value')]"
                ),
            },
            {
                # Fallback: second <em.value> on the page
                "kind": "xpath",
                "value": "(//em[contains(@class,'value')])[2]",
            },
        ]

    per_meta = check.get("metadata") or {}



    # Build metadata precedence: defaults -> metadata_by_url[url] -> per-check metadata
    defaults = cfg.get("defaults", {})
    meta_by_url = cfg.get("metadata_by_url", {})
    meta: Dict[str, str] = {
        "Site": defaults.get("site", ""),
        "Company": defaults.get("company", ""),
        "Service": "",
        "Menu": "",
    }
    url_meta = meta_by_url.get(url, {})
    for k, v in url_meta.items():
        meta[k.capitalize()] = v
    for k, v in per_meta.items():
        meta[k.capitalize()] = v

    print(f"\n→ {name} @ {url}")

    # Optional per-check login first (if provided in YAML)
    login_url = check.get("login_url") or LOGIN_URL
    login_user = check.get("login_username")
    login_pass = check.get("login_password")

    # If this is a Xertica host and no explicit overrides, prefer XERTICA_* env if present
    if (login_url and "xertica" in login_url.lower()) and (login_user is None and login_pass is None):
        login_user = os.getenv("XERTICA_USERNAME") or login_user
        login_pass = os.getenv("XERTICA_PASSWORD") or login_pass

    # Try logging in to login_url (if set)
    try:
        if login_url:
            open_with_sso(login_url, name.replace(" ", "_") + "_login",
                          username_override=login_user, password_override=login_pass)
    except Exception as e:
        print(f"Login warning for {login_url}: {e} (continuing)")

    # --- NEW: also prefer Xertica creds for the target URL navigation if it's an xertica domain ---
    nav_user = None
    nav_pass = None
    if "xertica" in (url or "").lower():
        nav_user = os.getenv("XERTICA_USERNAME") or login_user
        nav_pass = os.getenv("XERTICA_PASSWORD") or login_pass
    else:
        # fall back to whatever overrides were set (or None)
        nav_user = login_user
        nav_pass = login_pass

    # Open target URL
    try:
        open_with_sso(url, name.replace(" ", "_"),
                      username_override=nav_user, password_override=nav_pass)
    except Exception as e:
        print(f"Navigation warning for {url}: {e} (continuing)")

    # Xertica scheduler-specific pre-steps (company switch + light option select)
    try:
        if "asset.xertica.cloud/asset/scheduler" in (url or "").lower() or \
           ("asset.xertica.cloud/asset/scheduler" in driver.current_url.lower()):
            try:
                switched = switch_company_to_force("Xertica Clientes por reconocer")
                print("Company switch attempted:", switched)
            except Exception as e:
                print("switch_company_to_force error (continuing):", e)
            try:
                tried = select_only_xertica_option("Xertica Clientes por reconocer")
                print("Xertica option selection attempted:", tried)
            except Exception as e:
                print("select_only_xertica_option error (continuing):", e)
    except Exception as e:
        print("Xertica pre-step error:", e)

    # Let SPA render something
    try:
        WebDriverWait(driver, 60).until(
            lambda d: len(d.find_elements(By.CSS_SELECTOR,
                    "em.value, .value, .num, .number, .count, .am5-layer")) > 0
        )
    except TimeoutException:
        pass

    # Optional deep locator debug
    if LOCATOR_DEBUG and locators:
        try:
            for loc in locators:
                kind = (loc.get("kind") or "xpath").lower()
                value = (loc.get("value") or "").strip()
                if not value:
                    continue
                try:
                    elems = (driver.find_elements(By.XPATH, value)
                             if kind == "xpath" else driver.find_elements(By.CSS_SELECTOR, value))
                    print(f"DEBUG locator ({kind}): {value}  ->  matches: {len(elems)}")
                    for i, e in enumerate(elems[:5]):
                        try:
                            outer = driver.execute_script("return arguments[0].outerHTML.slice(0,200);", e)
                        except Exception:
                            outer = "<outerHTML unavailable>"
                        try:
                            selenium_text = e.text
                        except Exception:
                            selenium_text = ""
                        try:
                            raw_text = driver.execute_script("return arguments[0].textContent;", e) or ""
                            inner_text = driver.execute_script("return arguments[0].innerText;", e) or ""
                        except Exception:
                            raw_text = inner_text = ""
                        print(f"   - match[{i}] displayed={e.is_displayed()} text={repr(selenium_text)}")
                        print(f"     textContent={repr(raw_text)}, innerText={repr(inner_text)}")
                        print(f"     outerHTML~200={outer}")
                except Exception as ex:
                    print("  DEBUG locator error for", value, ":", ex)
        except Exception:
            pass

    # Try primary locators
    elem = None
    matched = None
    for loc in locators:
        kind = (loc.get("kind") or "xpath").lower()
        value = (loc.get("value") or "").strip()
        if not value:
            continue
        try:
            cond = EC.visibility_of_element_located(
                (By.XPATH, value) if kind == "xpath" else (By.CSS_SELECTOR, value)
            )
            elem = WebDriverWait(driver, 60).until(cond)
            matched = f"{kind}:{value}"
            break
        except TimeoutException:
            continue

    # Specialized types
    if ctype == "mtd_cost":
        value = wait_non_empty_text(lambda: elem.text, 30) if elem else ""
        if not value or "$" not in value:
            print("Falling back to JS scan for Month to Date Cost…")
            value = js_find_mtd_cost().strip()
        status = "PASS" if value else "FAIL"
        res = make_result(meta, url, name, matched or "[JS MTD cost]", value, status)

    elif ctype == "more_available_total":
        value = (elem.text.strip() if elem else "") or js_find_more_available_total().strip()
        status = "PASS" if value else "FAIL"
        res = make_result(meta, url, name, matched or "[JS MoreAvailable Total]", value, status)

    elif ctype == "cei_grade":
        value = (wait_non_empty_text(lambda: elem.text, 15) if elem else "") or js_find_ceikpi_grade_full().strip()
        status = "PASS" if value else "FAIL"
        res = make_result(meta, url, name, matched or "[JS CEI grade]", value, status)

    elif ctype == "element_exists":
        status = "PASS" if elem else "FAIL"
        value = "FOUND" if elem else ""
        if status == "FAIL" and js_fb.get("strategy") == "ec2_near_aws":
            print("Falling back to JS: EC2 near AWS…")
            if js_find_ec2_near_aws():
                status, value, matched = "PASS", "FOUND", "[JS EC2-near-AWS]"
        res = make_result(meta, url, name, matched or "n/a", value, status)

    else:  # value_required (generic)
        value = wait_non_empty_text(lambda: elem.text, 15) if elem else ""
        if (not value) and js_fb.get("strategy") == "scan_labels":
            print("Falling back to JS scan for label-based value…")
            rows = js_scan_labels()
            dump_json("Asset_ScanLabels", rows)
            label_keys = js_fb.get("label_keys") or SERVER_LABEL_KEYS
            value = pick_value_by_labels(rows, label_keys)
        status = "PASS" if value else "FAIL"
        res = make_result(meta, url, name, matched or "[JS scan fallback]", value, status)

    if res["Status"] == "FAIL":
        reason = f"{ctype} failed (locator: {res['Locator']})"
        screenshot_full = os.path.join("screenshots", res["Screenshot"]) if res["Screenshot"] else None
        slack_notify(name, reason, screenshot_full)

    print(f"Result: {res['Status']} | Value: {res['Value']}")
    
    if time.time() - start_time > MAX_CHECK_TIME:
        print(f"{name}: HARD TIMEOUT (>{MAX_CHECK_TIME}s)")
        return make_result(meta, url, name, "", "", "FAIL")
    
    return res
    




# ============================ MAIN ============================

if __name__ == "__main__":
    print("Starting health check…")
    try:
        ensure_dirs()
        # Create driver up front so helpers can use it
        driver, wait = create_driver()

        # Optionally fetch config.yaml from S3
        cfg_path = maybe_fetch_config_from_s3(CONFIG_URI, CONFIG_PATH_LOCAL)
        cfg = load_config(cfg_path)

        # Allow YAML to override a couple of runtime defaults if provided
        TIMEOUT_yaml = cfg.get("defaults", {}).get("timeout")
        RENDER_yaml  = cfg.get("defaults", {}).get("render_retry")
        if TIMEOUT_yaml:
            TIMEOUT = int(TIMEOUT_yaml)
        if RENDER_yaml:
            RENDER_RETRY = int(RENDER_yaml)

        # Initial console login (if LOGIN_URL is set)
        login_console()

        results: List[Dict[str, str]] = []

        for c in cfg.get("checks", []):
            try:
                results.append(run_one_check(c, cfg))

            except Exception as e:
                check_name = c.get("name", "<unnamed>")
                print(f"\n❌ Check crashed: {check_name}")
                print(f"   Reason: {e}")

                screenshot_path = None
                try:
                    screenshot_path = snap("Check_Crashed_" + safe_filename(check_name))
                    dump_html("Check_Crashed_" + safe_filename(check_name))
                except Exception as se:
                    print(f"   Failed to capture crash screenshot/html: {se}")

                meta = {"Site": "", "Company": "", "Service": "", "Menu": ""}
                row = make_result(
                    meta,
                    c.get("url", ""),
                    check_name,
                    "",
                    "",
                    "FAIL",
                    screenshot_path=screenshot_path,
                )

                # Slack for crash
                slack_notify(
                    check_name,
                    f"Check crashed with exception: {e}",
                    screenshot_path,
                )

                results.append(row)

        # Generate Excel + upload to S3 (and delete local file inside save_report)
        save_report(results)
        print("\nDone.")

    except Exception as e:
        # Top-level fatal error (before or outside checks loop)
        print(f"Fatal error: {e}")
        try:
            snap("Fatal_Error")
            dump_html("Fatal_Error")
        except Exception:
            pass
        slack_notify("Fatal Error", str(e), None)

    finally:
        print("Closing browser…")
        try:
            if driver:
                driver.quit()
        except Exception:
            pass
        # Clean up the temp Chrome profile if created
        try:
            if TEMP_PROFILE_DIR and os.path.isdir(TEMP_PROFILE_DIR):
                shutil.rmtree(TEMP_PROFILE_DIR, ignore_errors=True)
        except Exception:
            pass


